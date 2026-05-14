"""Export agent run results to a multi-sheet XLSX workbook.

Sheets produced:
  Run Info           — metadata (date, version, model, filters, run ID)
  Release Planning   — one row per epic (component, key, summary, status,
                       version, dev/qe/docs SP existing + proposed).
                       In summary-only mode the proposed SP columns are
                       omitted.
  QE & Docs Stories  — proposed QE/Docs stories with a blank Approved?
                       column for the epic owner to review.
                       Always present (headers only if no stories).
                       Omitted in summary-only mode.
  Observability Stories — proposed observability stories (metrics, alerts,
                       dashboards, telemetry) with a blank Approved?
                       column.  Always present (headers only if no
                       stories).  Omitted in summary-only mode.

Both story sheets include a Description column so the workbook is
self-contained and can be applied directly via --apply-xlsx without
requiring the companion JSON plan file.  ``read_xlsx_plan`` reads the
workbook and returns approved stories keyed by epic key.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from schemas.stories import StoryPayload

if TYPE_CHECKING:
    from agent.runner import _EpicTally, _RunCounters


# ── Palette ──────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F497D")   # dark blue
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")   # light blue
_BOLD        = Font(bold=True)
_CENTER      = Alignment(horizontal="center", vertical="top", wrap_text=True)
_TOP_LEFT    = Alignment(horizontal="left",   vertical="top", wrap_text=True)


def _header_row(ws: Any, cols: list[str]) -> None:
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _auto_width(ws: Any, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_w, max(min_w, length + 2))


def _freeze_top(ws: Any) -> None:
    ws.freeze_panes = "A2"


def _shade_alt_rows(ws: Any) -> None:
    """Shade every other data row (after the header)."""
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if i % 2 == 0:
            for cell in row:
                cell.fill = _ALT_FILL


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_run_info_sheet(
    ws: Any,
    metadata: dict[str, str],
) -> None:
    ws.title = "Run Info"
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    for key, val in metadata.items():
        ws.append([key, val])
    _auto_width(ws)
    ws.column_dimensions["B"].width = 60


def _build_summary_sheet(
    ws: Any,
    tallies: list[_EpicTally],
    jira_url: str = "",
    summary_only: bool = False,
    version: str = "",
) -> None:
    sheet_name = "Release Planning"
    if version and version != "(not set)":
        sheet_name = f"Release Planning {version}"
    ws.title = sheet_name

    has_components = any(t.components for t in tallies)
    _NOTABLE_LABELS = {"cnv-observability"}

    cols = ["Epic Key", "Summary", "Status"]
    if has_components:
        cols.append("Component")
    cols += ["Fix Version", "Target Version", "Labels"]
    if summary_only:
        cols += [
            "Dev SP (existing)",
            "QE SP (existing)",
            "Docs SP (existing)",
        ]
    else:
        cols += [
            "Dev SP (existing)", "Dev SP (proposed)",
            "QE SP (existing)", "QE SP (proposed)",
            "Docs SP (existing)", "Docs SP (proposed)",
            "Total Proposed SP",
        ]
    _header_row(ws, cols)

    for tally in tallies:
        notable = ", ".join(
            lb for lb in getattr(tally, "labels", [])
            if lb in _NOTABLE_LABELS
        )
        row: list[Any] = [tally.key, tally.summary, tally.status]
        if has_components:
            row.append(", ".join(tally.components))
        row += [
            tally.fix_version or "",
            tally.target_version or "",
            notable,
        ]
        if summary_only:
            row += [
                tally.dev_sp_existing,
                "no-qe" if tally.has_no_qe else tally.qe_sp_existing,
                "no-doc" if tally.has_no_doc else tally.docs_sp_existing,
            ]
        else:
            row += [
                tally.dev_sp_existing,
                tally.dev_sp_proposed,
                "no-qe" if tally.has_no_qe else tally.qe_sp_existing,
                "no-qe" if tally.has_no_qe else tally.qe_sp_proposed,
                "no-doc" if tally.has_no_doc else tally.docs_sp_existing,
                "no-doc" if tally.has_no_doc else tally.docs_sp_proposed,
                tally.dev_sp_proposed + (
                    0 if tally.has_no_qe else tally.qe_sp_proposed
                ) + (
                    0 if tally.has_no_doc else tally.docs_sp_proposed
                ),
            ]
        ws.append(row)

        # Hyperlink the Epic Key cell to Jira if URL is configured.
        if jira_url:
            cell = ws.cell(ws.max_row, 1)
            cell.hyperlink = (
                f"{jira_url.rstrip('/')}/browse/{tally.key}"
            )
            cell.font = Font(color="0563C1", underline="single")

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)


_QE_DOCS_CATEGORIES = {"qe", "docs"}
_OBS_CATEGORIES = {"metrics", "alerts", "dashboards", "telemetry"}


def _is_obs_sheet_story(story: StoryPayload) -> bool:
    """Return True if this story belongs on the Observability Stories sheet.

    Core observability categories always go there.  QE/Docs stories also
    go there when they cover a *proposed* observability item — indicated
    by a non-empty ``linked_to`` field set by the LLM.  QE/Docs stories
    without ``linked_to`` cover existing Jira child issues and go to the
    QE & Docs sheet instead.
    """
    if story.category in _OBS_CATEGORIES:
        return True
    if story.category in _QE_DOCS_CATEGORIES and story.linked_to:
        return True
    return False


def _build_review_sheet(
    ws: Any,
    title: str,
    plan_collector: dict[str, list[StoryPayload]],
    tallies: list[_EpicTally],
    obs_sheet: bool = False,
    jira_url: str = "",
) -> None:
    """Populate a review sheet for one story group.

    Args:
        title:        Sheet title.
        obs_sheet:    When True, include observability stories plus any
                      QE/Docs stories linked to proposed obs items.
                      When False, include only QE/Docs stories that cover
                      existing Jira child issues (no ``linked_to``).
        plan_collector: Full plan keyed by epic_key.
        tallies:      Used to look up epic summaries and components.
        jira_url:     Base Jira URL for hyperlinks (optional).
    """
    ws.title = title

    summary_map: dict[str, str] = {t.key: t.summary for t in tallies}
    comp_map: dict[str, str] = {
        t.key: ", ".join(t.components) for t in tallies
    }

    cols = [
        "Epic Key", "Epic Summary", "Component", "Category",
        "Story Summary", "Story Points", "Reasoning", "Description",
    ]
    if obs_sheet:
        cols.append("Covers proposed story")
    cols.append("Approved?")
    _header_row(ws, cols)

    for epic_key, stories in sorted(plan_collector.items()):
        epic_summary = summary_map.get(epic_key, "")
        component = comp_map.get(epic_key, "")
        for story in stories:
            belongs = (
                _is_obs_sheet_story(story) if obs_sheet
                else (
                    story.category in _QE_DOCS_CATEGORIES
                    and not story.linked_to
                )
            )
            if not belongs:
                continue
            row: list[Any] = [
                epic_key,
                epic_summary,
                component,
                story.category,
                story.summary,
                story.story_points or "",
                story.reasoning,
                story.description,
            ]
            if obs_sheet:
                row.append(story.linked_to or "")
            row.append("")  # Approved? — owner fills in
            ws.append(row)
            if jira_url:
                cell = ws.cell(ws.max_row, 1)
                cell.hyperlink = (
                    f"{jira_url.rstrip('/')}/browse/{epic_key}"
                )
                cell.font = Font(color="0563C1", underline="single")

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)
    # Reasoning and Description columns — cap width and wrap text.
    for col_name in ("Reasoning", "Description"):
        idx = cols.index(col_name) + 1
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 60
        for cell in ws[col_letter]:
            cell.alignment = _TOP_LEFT


# ── Public entry point ────────────────────────────────────────────────────────

def build_xlsx(
    path: str,
    *,
    metadata: dict[str, str],
    tallies: list[_EpicTally],
    plan_collector: dict[str, list[StoryPayload]],
    jira_url: str = "",
    summary_only: bool = False,
) -> None:
    """Write a multi-sheet XLSX workbook to *path*.

    Args:
        path:            Destination file path (e.g. ``report-20260511.xlsx``).
        metadata:        Key/value pairs for the Run Info sheet (date, model…).
        tallies:         Per-epic tally objects from ``_RunCounters``.
        plan_collector:  Proposed stories keyed by epic key.
        jira_url:        Base Jira URL for hyperlinks (optional).
        summary_only:    When True, omit proposed-SP columns from the Summary
                         sheet and skip the story review sheets entirely.  Use
                         with ``--summary-only`` to get a lightweight inventory
                         report without running the LLM.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    version = metadata.get("Version", "")
    _build_run_info_sheet(wb.create_sheet("Run Info"), metadata)
    _build_summary_sheet(
        wb.create_sheet("Release Planning"), tallies,
        jira_url=jira_url, summary_only=summary_only,
        version=version,
    )
    if not summary_only:
        _build_review_sheet(
            wb.create_sheet("QE & Docs Stories"),
            title="QE & Docs Stories",
            plan_collector=plan_collector,
            tallies=tallies,
            obs_sheet=False,
            jira_url=jira_url,
        )
        _build_review_sheet(
            wb.create_sheet("Observability Stories"),
            title="Observability Stories",
            plan_collector=plan_collector,
            tallies=tallies,
            obs_sheet=True,
            jira_url=jira_url,
        )

    wb.save(path)


# ── XLSX plan reader ──────────────────────────────────────────────────────────

_STORY_SHEET_NAMES = ("QE & Docs Stories", "Observability Stories")
_REJECTED_VALUES = {"no", "rejected", "reject", "n", "skip", "skipped"}


def read_xlsx_plan(
    path: str,
) -> tuple[str, dict[str, list[StoryPayload]]]:
    """Read approved stories from a report XLSX and return them keyed by epic.

    Reads both story sheets ("QE & Docs Stories" and "Observability Stories").
    A row is included only when its "Approved?" cell is non-empty and not a
    rejection value (no / rejected / skip / n).

    Returns:
        (version, plan_collector) where version is read from the Run Info
        sheet and plan_collector maps epic_key → list[StoryPayload].

    Raises:
        ValueError if required sheets or columns are missing.
    """
    wb = load_workbook(path, data_only=True)

    # Extract version from Run Info sheet.
    version = ""
    if "Run Info" in wb.sheetnames:
        for row in wb["Run Info"].iter_rows(min_row=2, values_only=True):
            if row and str(row[0] or "").strip() == "Version":
                version = str(row[1] or "").strip()
                break

    plan: dict[str, list[StoryPayload]] = {}

    for sheet_name in _STORY_SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [str(c.value or "").strip() for c in ws[1]]

        def _col(name: str) -> int | None:
            try:
                return headers.index(name)
            except ValueError:
                return None

        idx_epic    = _col("Epic Key")
        idx_cat     = _col("Category")
        idx_summary = _col("Story Summary")
        idx_sp      = _col("Story Points")
        idx_reason  = _col("Reasoning")
        idx_desc    = _col("Description")
        idx_linked  = _col("Covers proposed story")
        idx_approv  = _col("Approved?")

        if any(i is None for i in (
            idx_epic, idx_cat, idx_summary, idx_desc, idx_approv,
        )):
            missing = [
                n for n, i in [
                    ("Epic Key", idx_epic),
                    ("Category", idx_cat),
                    ("Story Summary", idx_summary),
                    ("Description", idx_desc),
                    ("Approved?", idx_approv),
                ]
                if i is None
            ]
            raise ValueError(
                f"Sheet '{sheet_name}' is missing required columns: "
                f"{', '.join(missing)}"
            )

        for row in ws.iter_rows(min_row=2, values_only=True):
            approved_raw = str(row[idx_approv] or "").strip().lower()
            if not approved_raw or approved_raw in _REJECTED_VALUES:
                continue

            epic_key = str(row[idx_epic] or "").strip()
            if not epic_key:
                continue

            try:
                sp = int(row[idx_sp]) if idx_sp is not None and row[idx_sp] is not None else None
            except (ValueError, TypeError):
                sp = None

            story = StoryPayload(
                category=str(row[idx_cat] or "").strip(),
                summary=str(row[idx_summary] or "").strip(),
                description=str(row[idx_desc] or "").strip(),
                story_points=sp,
                reasoning=str(row[idx_reason] or "").strip() if idx_reason is not None else "",
                linked_to=str(row[idx_linked] or "").strip() if idx_linked is not None else "",
            )
            plan.setdefault(epic_key, []).append(story)

    return version, plan
