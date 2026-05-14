"""Tests for agent/export/xlsx_report.py."""

import tempfile
import os

import openpyxl
import pytest

from agent.export.xlsx_report import build_xlsx, read_xlsx_plan
from agent.runner import _EpicTally
from schemas.stories import StoryPayload


def _make_tally(
    key: str,
    *,
    summary: str = "Epic summary",
    status: str = "groomed",
    components: list[str] | None = None,
    labels: list[str] | None = None,
    fix_version: str = "",
    target_version: str = "",
    dev_ex: int = 0, dev_pr: int = 0,
    qe_ex: int = 0, qe_pr: int = 0,
    docs_ex: int = 0, docs_pr: int = 0,
    has_no_qe: bool = False,
    has_no_doc: bool = False,
) -> _EpicTally:
    t = _EpicTally(key, status=status)
    t.summary = summary
    t.components = components or []
    t.labels = labels or []
    t.fix_version = fix_version
    t.target_version = target_version
    t.dev_sp_existing = dev_ex
    t.dev_sp_proposed = dev_pr
    t.qe_sp_existing = qe_ex
    t.qe_sp_proposed = qe_pr
    t.docs_sp_existing = docs_ex
    t.docs_sp_proposed = docs_pr
    t.has_no_qe = has_no_qe
    t.has_no_doc = has_no_doc
    return t


def _summary_ws(wb: openpyxl.Workbook):
    """Return the Release Planning sheet regardless of version suffix."""
    for name in wb.sheetnames:
        if name.startswith("Release Planning"):
            return wb[name]
    raise KeyError("No 'Release Planning' sheet found")


class TestBuildXlsx:
    """Verify the XLSX workbook structure and content."""

    def _build(self, tallies, plan_collector=None, metadata=None):
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False
        ) as f:
            path = f.name
        try:
            build_xlsx(
                path,
                metadata=metadata or {"Date": "2026-05-11", "Run ID": "abc"},
                tallies=tallies,
                plan_collector=plan_collector or {},
            )
            wb = openpyxl.load_workbook(path)
            return wb
        finally:
            os.unlink(path)

    def test_sheet_names_without_stories(self):
        wb = self._build([_make_tally("CNV-100")])
        assert "Run Info" in wb.sheetnames
        assert any(n.startswith("Release Planning") for n in wb.sheetnames)
        # Legacy "Stories" tab is gone; replaced by two review sheets.
        assert "Stories" not in wb.sheetnames
        assert "QE & Docs Stories" in wb.sheetnames
        assert "Observability Stories" in wb.sheetnames

    def test_summary_sheet_name_includes_version(self):
        wb = self._build(
            [_make_tally("CNV-100")],
            metadata={"Date": "2026-05-13", "Version": "5.0.0"},
        )
        assert "Release Planning 5.0.0" in wb.sheetnames

    def test_summary_sheet_name_without_version(self):
        wb = self._build(
            [_make_tally("CNV-100")],
            metadata={"Date": "2026-05-13"},
        )
        assert "Release Planning" in wb.sheetnames

    def test_review_sheets_always_present_when_not_summary_only(self):
        wb = self._build([_make_tally("CNV-100")])
        assert "QE & Docs Stories" in wb.sheetnames
        assert "Observability Stories" in wb.sheetnames

    def test_review_sheets_absent_in_summary_only_mode(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            build_xlsx(
                path,
                metadata={"Date": "2026-05-14"},
                tallies=[_make_tally("CNV-100")],
                plan_collector={},
                summary_only=True,
            )
            wb = openpyxl.load_workbook(path)
            assert "QE & Docs Stories" not in wb.sheetnames
            assert "Observability Stories" not in wb.sheetnames
        finally:
            os.unlink(path)

    def test_summary_sheet_has_epic_row(self):
        t = _make_tally(
            "CNV-200", summary="GPU metrics",
            dev_ex=5, dev_pr=3, qe_pr=2,
            components=["CNV Install"],
        )
        wb = self._build([t])
        ws = _summary_ws(wb)
        values = [
            str(cell.value or "") for row in ws.iter_rows(min_row=2)
            for cell in row
        ]
        assert "CNV-200" in values
        assert "GPU metrics" in values

    def test_run_info_sheet_has_metadata(self):
        wb = self._build(
            [_make_tally("CNV-300")],
            metadata={"Date": "2026-05-11", "Version": "5.0.0"},
        )
        ws = wb["Run Info"]
        values = [
            str(cell.value or "") for row in ws.iter_rows()
            for cell in row
        ]
        assert "Date" in values
        assert "2026-05-11" in values
        assert "Version" in values
        assert "5.0.0" in values

    def test_obs_stories_sheet_has_story_row(self):
        story = StoryPayload(
            category="alerts",
            summary="Add GPU alert",
            description="desc",
            story_points=2,
            reasoning="New alert needed",
        )
        wb = self._build(
            [_make_tally("CNV-400", summary="GPU passthrough epic",
                         components=["CNV Compute"])],
            plan_collector={"CNV-400": [story]},
        )
        ws = wb["Observability Stories"]
        values = [
            str(cell.value or "") for row in ws.iter_rows(min_row=2)
            for cell in row
        ]
        assert "CNV-400" in values
        assert "GPU passthrough epic" in values
        assert "Add GPU alert" in values
        assert "alerts" in values
        assert "New alert needed" in values

    def test_qe_docs_stories_sheet_has_story_row(self):
        story = StoryPayload(
            category="qe",
            summary="QE: test GPU passthrough",
            description="desc",
            story_points=3,
            reasoning="New feature needs test coverage",
        )
        wb = self._build(
            [_make_tally("CNV-401", summary="GPU passthrough epic")],
            plan_collector={"CNV-401": [story]},
        )
        ws = wb["QE & Docs Stories"]
        values = [
            str(cell.value or "") for row in ws.iter_rows(min_row=2)
            for cell in row
        ]
        assert "CNV-401" in values
        assert "QE: test GPU passthrough" in values
        assert "qe" in values

    def test_obs_stories_and_linked_qe_routed_to_obs_sheet(self):
        """Obs stories and QE stories linked to them land in Observability sheet."""
        obs_story = StoryPayload(
            category="metrics", summary="Add metric X",
            description="desc", story_points=2, reasoning="r",
        )
        qe_linked = StoryPayload(
            category="qe", summary="QE: verify metric X",
            description="desc", story_points=2, reasoning="r",
            linked_to="Add metric X",
        )
        wb = self._build(
            [_make_tally("CNV-402")],
            plan_collector={"CNV-402": [obs_story, qe_linked]},
        )
        obs_values = [
            str(c.value or "")
            for row in wb["Observability Stories"].iter_rows(min_row=2)
            for c in row
        ]
        qe_values = [
            str(c.value or "")
            for row in wb["QE & Docs Stories"].iter_rows(min_row=2)
            for c in row
        ]
        assert "Add metric X" in obs_values
        assert "QE: verify metric X" in obs_values, (
            "Linked QE story must appear in Observability sheet"
        )
        assert "QE: verify metric X" not in qe_values, (
            "Linked QE story must NOT appear in QE & Docs sheet"
        )

    def test_unlinked_qe_stays_in_qe_docs_sheet(self):
        """QE stories without linked_to (covering existing child) go to QE & Docs."""
        qe_unlinked = StoryPayload(
            category="qe", summary="QE: verify snapshot restore",
            description="desc", story_points=3, reasoning="r",
        )
        wb = self._build(
            [_make_tally("CNV-403")],
            plan_collector={"CNV-403": [qe_unlinked]},
        )
        qe_values = [
            str(c.value or "")
            for row in wb["QE & Docs Stories"].iter_rows(min_row=2)
            for c in row
        ]
        obs_values = [
            str(c.value or "")
            for row in wb["Observability Stories"].iter_rows(min_row=2)
            for c in row
        ]
        assert "QE: verify snapshot restore" in qe_values
        assert "QE: verify snapshot restore" not in obs_values

    def test_obs_sheet_has_covers_proposed_story_column(self):
        """Observability sheet includes a 'Covers proposed story' column."""
        wb = self._build([_make_tally("CNV-410")])
        headers = [
            str(cell.value or "") for cell in wb["Observability Stories"][1]
        ]
        assert "Covers proposed story" in headers

    def test_qe_docs_sheet_has_no_covers_proposed_story_column(self):
        """QE & Docs sheet does NOT include 'Covers proposed story' column."""
        wb = self._build([_make_tally("CNV-410")])
        headers = [
            str(cell.value or "") for cell in wb["QE & Docs Stories"][1]
        ]
        assert "Covers proposed story" not in headers

    def test_review_sheets_have_approved_column(self):
        wb = self._build([_make_tally("CNV-410")])
        for sheet_name in ("QE & Docs Stories", "Observability Stories"):
            headers = [
                str(cell.value or "") for cell in wb[sheet_name][1]
            ]
            assert "Approved?" in headers, (
                f"'Approved?' missing from {sheet_name}"
            )

    def test_review_sheets_have_epic_summary_column_header(self):
        wb = self._build([_make_tally("CNV-410")])
        for sheet_name in ("QE & Docs Stories", "Observability Stories"):
            headers = [
                str(cell.value or "") for cell in wb[sheet_name][1]
            ]
            assert "Epic Summary" in headers, (
                f"'Epic Summary' missing from {sheet_name}"
            )

    def test_linked_to_value_shown_in_obs_sheet(self):
        """linked_to value appears in the 'Covers proposed story' column."""
        qe_linked = StoryPayload(
            category="qe", summary="QE: verify dashboard",
            description="desc", story_points=2, reasoning="r",
            linked_to="Add Autopilot Health dashboard",
        )
        wb = self._build(
            [_make_tally("CNV-420")],
            plan_collector={"CNV-420": [qe_linked]},
        )
        ws = wb["Observability Stories"]
        all_values = [
            str(c.value or "")
            for row in ws.iter_rows(min_row=2) for c in row
        ]
        assert "Add Autopilot Health dashboard" in all_values

    def test_summary_only_omits_proposed_columns(self):
        t = _make_tally("CNV-500", dev_ex=5, dev_pr=3, qe_pr=2)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from agent.export.xlsx_report import build_xlsx
            build_xlsx(
                path,
                metadata={"Date": "2026-05-13"},
                tallies=[t],
                plan_collector={},
                summary_only=True,
            )
            wb = openpyxl.load_workbook(path)
            ws = _summary_ws(wb)
            headers = [str(cell.value or "") for cell in ws[1]]
            assert "Dev SP (existing)" in headers
            assert "Dev SP (proposed)" not in headers
            assert "Total Proposed SP" not in headers
            assert "QE & Docs Stories" not in wb.sheetnames
            assert "Observability Stories" not in wb.sheetnames
        finally:
            os.unlink(path)

    def test_no_qe_shown_in_summary(self):
        t = _make_tally("CNV-500", has_no_qe=True)
        wb = self._build([t])
        ws = _summary_ws(wb)
        values = [
            str(cell.value or "") for row in ws.iter_rows(min_row=2)
            for cell in row
        ]
        assert "no-qe" in values

    def test_component_column_present_when_components_set(self):
        t = _make_tally("CNV-600", components=["CNV Storage"])
        wb = self._build([t])
        ws = _summary_ws(wb)
        headers = [str(cell.value or "") for cell in ws[1]]
        assert "Component" in headers

    def test_labels_column_always_present(self):
        t = _make_tally("CNV-700")
        wb = self._build([t])
        ws = _summary_ws(wb)
        headers = [str(cell.value or "") for cell in ws[1]]
        assert "Labels" in headers

    def test_cnv_observability_label_shown_in_labels_column(self):
        t = _make_tally(
            "CNV-800",
            labels=["cnv-observability", "cnv-grooming-agent"],
        )
        wb = self._build([t])
        ws = _summary_ws(wb)
        values = [
            str(cell.value or "") for row in ws.iter_rows(min_row=2)
            for cell in row
        ]
        assert "cnv-observability" in values

    def test_non_notable_labels_not_shown(self):
        t = _make_tally("CNV-900", labels=["cnv-grooming-agent", "some-other"])
        wb = self._build([t])
        ws = _summary_ws(wb)
        values = [
            str(cell.value or "") for row in ws.iter_rows(min_row=2)
            for cell in row
        ]
        assert "cnv-grooming-agent" not in values
        assert "some-other" not in values

    def test_review_sheets_have_description_column(self):
        wb = self._build([_make_tally("CNV-910")])
        for sheet_name in ("QE & Docs Stories", "Observability Stories"):
            headers = [str(c.value or "") for c in wb[sheet_name][1]]
            assert "Description" in headers, (
                f"'Description' missing from {sheet_name}"
            )

    def test_description_value_in_story_row(self):
        story = StoryPayload(
            category="metrics",
            summary="Add metric X",
            description="Full story body text here.",
            story_points=3,
            reasoning="r",
        )
        wb = self._build(
            [_make_tally("CNV-920")],
            plan_collector={"CNV-920": [story]},
        )
        ws = wb["Observability Stories"]
        all_values = [
            str(c.value or "")
            for row in ws.iter_rows(min_row=2) for c in row
        ]
        assert "Full story body text here." in all_values


class TestReadXlsxPlan:
    """Tests for read_xlsx_plan — the XLSX → StoryPayload reader."""

    def _write_xlsx(
        self,
        stories: dict[str, list[StoryPayload]],
        version: str = "5.0.0",
        approved: dict[str, str] | None = None,
    ) -> str:
        """Build an XLSX and return its path. approved maps story summary → value."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        build_xlsx(
            path,
            metadata={"Date": "2026-05-14", "Version": version},
            tallies=[_make_tally(k) for k in stories],
            plan_collector=stories,
        )
        # Fill in Approved? values via openpyxl.
        wb = openpyxl.load_workbook(path)
        for sheet_name in ("QE & Docs Stories", "Observability Stories"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [str(c.value or "") for c in ws[1]]
            try:
                sum_idx = headers.index("Story Summary") + 1
                app_idx = headers.index("Approved?") + 1
            except ValueError:
                continue
            for row in ws.iter_rows(min_row=2):
                summary = str(row[sum_idx - 1].value or "")
                if approved and summary in approved:
                    row[app_idx - 1].value = approved[summary]
        wb.save(path)
        return path

    def test_returns_version(self):
        path = self._write_xlsx({})
        try:
            version, _ = read_xlsx_plan(path)
            assert version == "5.0.0"
        finally:
            os.unlink(path)

    def test_empty_approved_returns_nothing(self):
        story = StoryPayload(
            category="metrics", summary="Add metric X",
            description="desc", story_points=3, reasoning="r",
        )
        path = self._write_xlsx({"CNV-100": [story]})
        try:
            _, plan = read_xlsx_plan(path)
            assert plan == {}
        finally:
            os.unlink(path)

    def test_approved_story_included(self):
        story = StoryPayload(
            category="metrics", summary="Add metric X",
            description="Full description.", story_points=3, reasoning="r",
        )
        path = self._write_xlsx(
            {"CNV-100": [story]},
            approved={"Add metric X": "Yes"},
        )
        try:
            _, plan = read_xlsx_plan(path)
            assert "CNV-100" in plan
            assert len(plan["CNV-100"]) == 1
            assert plan["CNV-100"][0].category == "metrics"
            assert plan["CNV-100"][0].description == "Full description."
        finally:
            os.unlink(path)

    def test_rejected_story_excluded(self):
        story = StoryPayload(
            category="metrics", summary="Add metric X",
            description="desc", story_points=3, reasoning="r",
        )
        path = self._write_xlsx(
            {"CNV-100": [story]},
            approved={"Add metric X": "No"},
        )
        try:
            _, plan = read_xlsx_plan(path)
            assert plan == {}
        finally:
            os.unlink(path)

    def test_partial_approval(self):
        obs = StoryPayload(
            category="metrics", summary="Add metric X",
            description="desc", story_points=3, reasoning="r",
        )
        qe = StoryPayload(
            category="qe", summary="QE: verify snapshot",
            description="qe desc", story_points=2, reasoning="r",
        )
        path = self._write_xlsx(
            {"CNV-200": [obs, qe]},
            approved={"Add metric X": "Yes"},
        )
        try:
            _, plan = read_xlsx_plan(path)
            summaries = [s.summary for s in plan.get("CNV-200", [])]
            assert any("metric X" in s for s in summaries)
            assert not any("snapshot" in s for s in summaries)
        finally:
            os.unlink(path)

    def test_linked_qe_story_round_trips(self):
        """A QE story with linked_to set goes into the Observability sheet
        and is recovered by read_xlsx_plan with linked_to preserved."""
        qe = StoryPayload(
            category="qe", summary="QE: verify metric X",
            description="qe desc", story_points=2, reasoning="r",
            linked_to="Add metric X",
        )
        path = self._write_xlsx(
            {"CNV-300": [qe]},
            approved={"QE: verify metric X": "Yes"},
        )
        try:
            _, plan = read_xlsx_plan(path)
            stories = plan.get("CNV-300", [])
            assert stories, "Expected approved linked QE story"
            assert stories[0].linked_to == "Add metric X"
        finally:
            os.unlink(path)
